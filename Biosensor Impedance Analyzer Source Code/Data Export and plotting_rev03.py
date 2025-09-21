import serial
import threading
import queue
import time
import sys
import os
from prompt_toolkit import prompt
from prompt_toolkit.patch_stdout import patch_stdout
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import re
from matplotlib import font_manager, rc
from matplotlib.ticker import ScalarFormatter
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# ------------------------
# 0) 폰트/시리얼 설정 등
# ------------------------

# 한글 폰트 설정
font_path = 'C:/Windows/Fonts/malgun.ttf'  # 맑은 고딕 폰트 경로
try:
    font_prop = font_manager.FontProperties(fname=font_path)
    rc('font', family=font_prop.get_name())
except Exception as e:
    print(f"폰트 설정 오류: {e}")
    rc('font', family='DejaVu Sans')  # 기본 폰트로 설정

# 음수 기호 제대로 표시
plt.rcParams['axes.unicode_minus'] = False

# 시리얼 포트 설정
try:
    ser = serial.Serial('COM3', 115200, timeout=0.1)
    time.sleep(0.1)  # 시리얼 포트 안정화 시간
    ser.reset_input_buffer()
    ser.reset_output_buffer()
    print("시리얼 포트에 연결되었습니다.")
except serial.SerialException as e:
    print(f"시리얼 포트 오류: {e}")
    sys.exit(1)

# ------------------------
# 1) 글로벌 변수/엑셀 초기화
# ------------------------

prompt_queue = queue.Queue()

calibration_runs = []  # 캘리브레이션 런 정보 저장
calibration_data = []  # 캘리브레이션 데이터 저장

# 단일 스윕(모드 1, 2) 전용 리스트
measurement_data = []

# 범위 스윕(모드 4) 전용 리스트
range_data = []

current_mode = None
xAddrStr = ""
yAddrStr = ""
group_selected = None
calibration_impedance = ""
measurement_type = None

# 좌표 처리 변수
currentCoord = None  # 튜플 (X, Y)
next_x = None
next_y = None

# 이벤트 (범위 스윕 완료 시 플롯)
range_sweep_complete = threading.Event()

# 엑셀 파일 저장 관련
save_directory = "C:/Users/Hyunseo/OneDrive/Desktop/Data"
base_filename = "measurement_data"
file_extension = "xlsx"

def get_unique_filename(directory, base_name, extension):
    filename = f"{base_name}.{extension}"
    counter = 2
    while os.path.exists(os.path.join(directory, filename)):
        filename = f"{base_name}_{counter}.{extension}"
        counter += 1
    return os.path.join(directory, filename)

excel_filename = get_unique_filename(save_directory, base_filename, file_extension)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Measurement Data"

current_calibration_run = 0
is_calibrating = False

def initialize_new_calibration_run(run_number):
    global current_calibration_run
    run_number = current_calibration_run
    start_col = 1 + 8 * run_number
    start_col_letter = get_column_letter(start_col)
    
    start_row = 1
    ws.cell(row=start_row, column=start_col, value="설정된 Calibration Impedance : ")
    
    # 기존 캘리브레이션 헤더
    cal_headers = ['Cal Point', 'R / I', '|Z|', 'System Phase']
    for i, header in enumerate(cal_headers):
        ws.cell(row=start_row + 1, column=start_col + i, value=header)
    
    wb.save(excel_filename)
    print(f"캘리브레이션 런 {run_number} 초기화 완료. 시작 열: {start_col_letter}")
    calibration_runs.append({'run_number': run_number, 'start_col': start_col, 'data': []})

initialize_new_calibration_run(current_calibration_run)
print(f"엑셀 파일 '{excel_filename}'이(가) 생성되었습니다.")

# 단일 스윕 완료 신호
sweep_complete = threading.Event()

def is_prompt_line(line):
    prompts = [
        "시작 주파수를 입력하세요",
        "주파수 증가량을 입력하세요",
        "측정 횟수를 입력하세요",
        "Settling Time Cycles를 입력하세요",
        "Output Excitation Range를 선택하세요",
        "PGA Gain을 선택하세요",
        "Calibration Impedance를 입력하세요",
        "MUX 그룹을 선택하세요",
        "X Axis Address",
        "Y Axis Address",
        "AD5933 모드 설정",
        "Bit",
        "이 범위가 맞습니까? (Y/N)"
    ]
    return any(line.strip().startswith(prompt) for prompt in prompts)

def parse_calibration_line(line):
    try:
        pattern = r"Cal Point (\d+):\s+R=(-?\d+) / I=(-?\d+)\s+\|Z\|=([\d.]+)\s+System Phase=([\d.+-]+) degrees"
        match = re.match(pattern, line)
        if match:
            cal_point = f"Cal Point {match.group(1)}"
            r_i = f"R={match.group(2)} / I={match.group(3)}"
            z = match.group(4)
            phase = f"{match.group(5)} degrees"
            return [cal_point, r_i, z, phase]
        else:
            return None
    except Exception as e:
        print(f"Calibration 데이터 파싱 오류: {e} - 라인: {line}")
        return None

def parse_measurement_line(line):
    try:
        pattern = (
            r"(\d+\.\d+)kHz:\s+R=(-?\d+)/I=(-?\d+)\s+\|Z\|=([-+]?\d+\.\d+)\s+"
            r"Phase=([-+]?\d+\.\d+) degrees\s+Resistance=([-+]?\d+\.\d+)\s+"
            r"Reactance=([-+]?\d+\.\d+)"
        )
        match = re.match(pattern, line)
        if match:
            freq_khz = float(match.group(1))
            freq = f"{int(freq_khz * 1000)} Hz"
            r_i = f"R={match.group(2)} / I={match.group(3)}"
            impedance = float(match.group(4))
            phase = float(match.group(5))
            resistance = float(match.group(6))
            reactance = float(match.group(7))
            return [freq, r_i, impedance, phase, resistance, reactance]
        else:
            return None
    except (IndexError, ValueError) as e:
        print(f"Measurement 데이터 파싱 오류: {e} - 라인: {line}")
        return None

def add_headers(current_run, headers):
    start_col = current_run['start_col']
    if 'current_row' not in current_run:
        current_run['current_row'] = 1
    
    header_row = current_run['current_row']
    
    for i, header in enumerate(headers):
        cell = ws.cell(row=header_row, column=start_col + i, value=header)
        
        # 글씨를 굵게
        cell.font = Font(bold=True)
        
        # 노란색 배경
        cell.fill = PatternFill(start_color='FFFF00', 
                                end_color='FFFF00', 
                                fill_type='solid')
        
        # 셀 가운데 정렬
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_run['current_row'] += 1
    wb.save(excel_filename)
    print(f"헤더 추가: {headers}")

HEADER_FIELDS = [
    'Freq (Hz)', 
    'R / I', 
    '|Z|', 
    'Phase (Degrees)', 
    'Resistance', 
    'Reactance', 
    'X', 
    'Y'
]

def read_from_port():
    global current_mode, xAddrStr, yAddrStr, calibration_impedance, group_selected
    global measurement_type, current_calibration_run, is_calibrating
    global currentCoord, next_x, next_y

    while True:
        if not ser.is_open:
            print("시리얼 포트가 닫혀있습니다. 데이터 수신 루프 종료.")
            break
        try:
            line = ser.readline().decode('utf-8', errors='ignore').strip()
            if line:
                # 1) 제어/캘리브레이션 관련 처리
                if "캘리브레이션을 시작합니다." in line:
                    if not is_calibrating:
                        print("\n캘리브레이션을 시작합니다. 새로운 캘리브레이션 런을 초기화합니다.\n")
                        current_calibration_run += 1
                        initialize_new_calibration_run(current_calibration_run)
                        is_calibrating = False
                        ser.reset_input_buffer()
                        ser.reset_output_buffer()
                    continue

                if "ESP-ROM" in line:
                    print("\n디바이스가 리셋되었습니다. 새로운 캘리브레이션 런을 시작합니다.\n")
                    current_calibration_run += 1
                    initialize_new_calibration_run(current_calibration_run)
                    ser.reset_input_buffer()
                    ser.reset_output_buffer()
                    continue

                if is_prompt_line(line):
                    prompt_queue.put(line)
                    continue

                if "설정된 Calibration Impedance" in line:
                    parts = line.split(':')
                    if len(parts) >= 2:
                        calibration_impedance = parts[1].strip().split(' ')[0]
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            ws.cell(row=1, column=start_col,
                                    value=f"설정된 Calibration Impedance : {calibration_impedance} ohm")
                            wb.save(excel_filename)
                            print(line)
                    continue

                if line.startswith("Cal Point"):
                    cal_data = parse_calibration_line(line)
                    if cal_data and calibration_runs:
                        calibration_data.append(cal_data)
                        current_run = calibration_runs[-1]
                        start_col = current_run['start_col']
                        if 'current_row' not in current_run:
                            current_run['current_row'] = 3
                        for i, data in enumerate(cal_data):
                            ws.cell(row=current_run['current_row'], column=start_col + i, value=data)
                        current_run['current_row'] += 1
                        wb.save(excel_filename)
                        print("\t".join(cal_data))
                    continue

                # 2) X/Y 주소 설정 (단일 스윕 시 Coord로 바로 설정)
                if "설정된 X축 Address" in line and "Y축 Address" in line:
                    parts = line.split(',')
                    if len(parts) >= 2:
                        xAddrStr = parts[0].split(':')[-1].strip()
                        yAddrStr = parts[1].split(':')[-1].strip()
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            if 'current_row' not in current_run:
                                current_run['current_row'] = 3
                            row_num = current_run['current_row']
                            # "설정된 좌표" | "X=..." | "Y=..." 로 저장
                            ws.cell(row=row_num, column=start_col, value="설정된 좌표")
                            ws.cell(row=row_num, column=start_col + 1, value=f"X={xAddrStr}")
                            ws.cell(row=row_num, column=start_col + 2, value=f"Y={yAddrStr}")
                            current_run['current_row'] += 1
                            wb.save(excel_filename)
                            print(line)
                            # 단일 스윕 모드(COB, Rcal)라면 바로 currentCoord로 저장
                            if measurement_type in ['COB', 'Rcal']:
                                currentCoord = (xAddrStr, yAddrStr)
                    continue

                # 3) Rcal, COB 모드 메시지
                if "Rcal 위치의 임피던스를 체크합니다." in line:
                    if calibration_runs:
                        current_run = calibration_runs[-1]
                        start_col = current_run['start_col']
                        if 'current_row' not in current_run:
                            current_run['current_row'] = 1
                        current_run['current_row'] += 1
                        ws.cell(row=current_run['current_row'], column=start_col, value="Rcal 위치의 임피던스를 체크합니다.")
                        current_run['current_row'] += 1
                        wb.save(excel_filename)
                        print(line)
                        measurement_type = 'Rcal'
                        add_headers(current_run, HEADER_FIELDS)
                    continue

                if "COB의 임피던스를 체크합니다." in line:
                    if calibration_runs:
                        current_run = calibration_runs[-1]
                        start_col = current_run['start_col']
                        if 'current_row' not in current_run:
                            current_run['current_row'] = 1
                        current_run['current_row'] += 1
                        ws.cell(row=current_run['current_row'], column=start_col, value="COB 위치의 임피던스를 체크합니다.")
                        current_run['current_row'] += 1
                        wb.save(excel_filename)
                        print(line)
                        measurement_type = 'COB'
                    continue

                # 4) 범위 스윕 관련 처리
                if "COB 범위 스윕 (7비트 입력 방식)을 시작합니다." in line:
                    if calibration_runs:
                        current_run = calibration_runs[-1]
                        start_col = current_run['start_col']
                        if 'current_row' not in current_run:
                            current_run['current_row'] = 3
                        current_run['current_row'] += 1
                        ws.cell(row=current_run['current_row'], column=start_col,
                                value="COB 범위 스윕 (7비트 입력 방식)을 시작합니다.")
                        current_run['current_row'] += 1
                        wb.save(excel_filename)
                        print(line)
                        measurement_type = 'COB-range'
                    continue

                if "그룹" in line and "선택" in line:
                    match = re.search(r"그룹\s+(\d+)\s+선택", line)
                    if match:
                        group_selected = match.group(1)
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            if 'current_row' not in current_run:
                                current_run['current_row'] = 3
                            row_num = current_run['current_row']
                            ws.cell(row=row_num, column=start_col, value=f"그룹 {group_selected} 선택")
                            current_run['current_row'] += 1
                            wb.save(excel_filename)
                            print(line)
                            add_headers(current_run, HEADER_FIELDS)
                    continue

                # 5) 좌표 수신 처리 (범위 스윕에서는 별도로 좌표가 나오므로)
                match_x = re.search(r"현재 좌표:\s*X\s*=\s*(\d+)", line)
                if match_x:
                    next_x = match_x.group(1)
                    print(line)
                    continue

                match_y = re.search(r"Y\s*=\s*(\d+)", line)
                if match_y:
                    next_y = match_y.group(1)
                    print(line)
                    if next_x is not None and calibration_runs:
                        current_run = calibration_runs[-1]
                        start_col = current_run['start_col']
                        if 'current_row' not in current_run:
                            current_run['current_row'] = 3
                        current_run['current_row'] += 1
                        ws.cell(row=current_run['current_row'], column=start_col, value="현재 좌표")
                        ws.cell(row=current_run['current_row'], column=start_col + 1, value=f"X={next_x}")
                        ws.cell(row=current_run['current_row'], column=start_col + 2, value=f"Y={next_y}")
                        current_run['current_row'] += 1
                        wb.save(excel_filename)
                        # 범위 스윕에서는 수신된 좌표를 currentCoord로 설정
                        currentCoord = (next_x, next_y)
                        next_x = None
                        next_y = None
                    continue

                if "Frequency sweep complete!" in line:
                    print(line)
                    sweep_complete.set()
                    continue

                if "COB 범위 스윕 완료." in line:
                    print(line)
                    range_sweep_complete.set()
                    continue

                # 6) 실제 측정 데이터 처리
                print(line)
                data = parse_measurement_line(line)
                if data:
                    if measurement_type in ['COB', 'Rcal']:
                        # 단일 스윕 데이터 저장
                        if currentCoord:
                            data.append(currentCoord[0])  # X값
                            data.append(currentCoord[1])  # Y값
                        else:
                            data.append("N/A")
                            data.append("N/A")
                        measurement_data.append(data)
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            if 'current_row' not in current_run:
                                current_run['current_row'] = 3
                            measurement_row = current_run['current_row']
                            for i, datum in enumerate(data):
                                ws.cell(row=measurement_row, column=start_col + i, value=datum)
                            current_run['current_row'] += 1
                            wb.save(excel_filename)
                    elif measurement_type == 'COB-range':
                        # 범위 스윕 데이터 저장
                        if currentCoord:
                            data.append(currentCoord[0])
                            data.append(currentCoord[1])
                        else:
                            data.append("N/A")
                            data.append("N/A")
                        range_data.append(data)
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            if 'current_row' not in current_run:
                                current_run['current_row'] = 3
                            measurement_row = current_run['current_row']
                            for i, datum in enumerate(data):
                                ws.cell(row=measurement_row, column=start_col + i, value=datum)
                            current_run['current_row'] += 1
                            wb.save(excel_filename)
                    else:
                        # Unknown 모드 혹은 캘리브레이션 등은 무시
                        pass

        except serial.SerialException as e:
            if "액세스가 거부되었습니다" in str(e):
                print("시리얼 포트에 액세스가 거부되었습니다. 연결 해제된 것으로 판단하여 데이터 수신 루프 종료.")
                break
            else:
                print(f"데이터 수신 중 시리얼 포트 오류 발생: {e}")
                time.sleep(0.05)
        except Exception as e:
            print(f"데이터 수신 중 오류 발생: {e}")
            time.sleep(0.01)

def plot_data(data, mode_label=""):
    # 헤더 8개: X, Y 분리
    columns = [
        'freq (Hz)', 'R / I', '|Z|', 'Phase (Degrees)', 
        'Resistance', 'Reactance', 'X', 'Y'
    ]
    df = pd.DataFrame(data, columns=columns)
    
    if df.empty:
        print("플로팅할 데이터가 없습니다.")
        return

    df['Frequency'] = df['freq (Hz)'].apply(lambda x: int(x.replace(' Hz','')))

    # R, I 추출 (필요시)
    def extract_r_i(r_i_str):
        try:
            r, i = r_i_str.split('/')
            r = float(r.replace('R=','').strip())
            i = float(i.replace('I=','').strip())
            return r, i
        except:
            return None, None

    df[['R', 'I']] = df['R / I'].apply(lambda x: pd.Series(extract_r_i(x)))
    df = df.dropna(subset=['R','I'])

    # 임시로 Coord 문자열을 만들어서 그룹핑 (X,Y가 같으면 같은 그룹)
    df['Coord'] = df.apply(lambda row: f"X={row['X']},Y={row['Y']}", axis=1)

    if mode_label == '2':
        suptitle = "Rcal 위치 임피던스 측정 결과"
    elif mode_label == '1':
        suptitle = "COB 임피던스 측정 결과"
    elif mode_label == '4':
        suptitle = "COB 범위 스윕 결과"
    else:
        suptitle = "임피던스 측정 결과"

    color_cycle = plt.colormaps.get_cmap('tab10')
    groups = df.groupby('Coord')

    plt.figure(figsize=(15, 10))

    # subplot 1: Frequency vs R, I
    ax1 = plt.subplot(2,2,1)
    for idx, (coord, group) in enumerate(groups):
        color = color_cycle(idx % 10)
        ax1.scatter(group['Frequency'], group['R'], color=color, marker='o', label=f'R ({coord})')
        ax1.scatter(group['Frequency'], group['I'], color=color, marker='x', label=f'I ({coord})')
    ax1.set_title('Frequency vs R and I')
    ax1.set_xlabel('Frequency (Hz)')
    ax1.set_ylabel('R and I')
    ax1.legend()
    ax1.grid(True)

    # subplot 2: Frequency vs |Z|
    ax2 = plt.subplot(2,2,2)
    for idx, (coord, group) in enumerate(groups):
        color = color_cycle(idx % 10)
        ax2.scatter(group['Frequency'], group['|Z|'], color=color, marker='s', label=coord)
    ax2.set_title('Frequency vs |Z|')
    ax2.set_xlabel('Frequency (Hz)')
    ax2.set_ylabel('|Z| (Ohm)')
    ax2.set_yscale('linear')
    ax2.yaxis.set_major_formatter(ScalarFormatter(useOffset=False))
    ax2.grid(True)
    ax2.legend()

    # subplot 3: Frequency vs Phase (Degrees)
    ax3 = plt.subplot(2,2,3)
    for idx, (coord, group) in enumerate(groups):
        color = color_cycle(idx % 10)
        ax3.scatter(group['Frequency'], group['Phase (Degrees)'], color=color, marker='^', label=coord)
    ax3.set_title('Frequency vs Phase (Degrees)')
    ax3.set_xlabel('Frequency (Hz)')
    ax3.set_ylabel('Phase (Degrees)')
    ax3.set_ylim(-180, 180)
    ax3.grid(True)
    ax3.legend()

    # subplot 4: Frequency vs Resistance & Reactance
    ax4 = plt.subplot(2,2,4)
    for idx, (coord, group) in enumerate(groups):
        color = color_cycle(idx % 10)
        ax4.scatter(group['Frequency'], group['Resistance'], color=color, marker='D', label=f'Res ({coord})')
        ax4.scatter(group['Frequency'], group['Reactance'], color=color, marker='v', label=f'React ({coord})')
    ax4.set_title('Frequency vs Resistance & Reactance')
    ax4.set_xlabel('Frequency (Hz)')
    ax4.set_ylabel('Value')
    ax4.grid(True)
    ax4.legend()

    plt.suptitle(suptitle, fontsize=16)
    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    plt.show()

thread = threading.Thread(target=read_from_port, daemon=True)
thread.start()

try:
    while True:
        if not prompt_queue.empty():
            prompt_text = prompt_queue.get()
            with patch_stdout():
                user_input = prompt(prompt_text + ' ')
            try:
                ser.write((user_input.strip() + '\n').encode('utf-8'))
                if "AD5933 모드 설정" in prompt_text:
                    current_mode = user_input.strip()
                    if current_mode == '1':
                        measurement_type = 'COB'
                    elif current_mode == '2':
                        measurement_type = 'Rcal'
                    elif current_mode == '3':
                        measurement_type = 'COB-diagonal'
                    elif current_mode == '4':
                        measurement_type = 'COB-range'
                    elif current_mode == '0':
                        is_calibrating = True
                        current_calibration_run += 1
                        initialize_new_calibration_run(current_calibration_run)
                        print(f"캘리브레이션 런 {current_calibration_run}으로 이동했습니다.")
                    else:
                        measurement_type = 'Unknown'
            except serial.SerialException as e:
                print(f"시리얼 포트로 데이터 전송 중 오류 발생: {e}")
                break

        elif sweep_complete.is_set():
            if current_mode in ['1', '2']:
                print("\n[INFO] 단일 스윕 완료 - 현재 스윕 데이터 플롯 (COB/Rcal)\n")
                plot_data(measurement_data, mode_label=current_mode)
                measurement_data.clear()
                currentCoord = None
            sweep_complete.clear()

        elif range_sweep_complete.is_set():
            print("\n[INFO] COB 범위 스윕이 완료되어 범위 스윕 데이터를 플롯합니다.\n")
            plot_data(range_data, mode_label='4')
            range_data.clear()
            currentCoord = None
            range_sweep_complete.clear()
        else:
            time.sleep(0.01)

except KeyboardInterrupt:
    print("\n프로그램을 종료합니다.")
    if measurement_data:
        plot_data(measurement_data, mode_label=current_mode)
        print(f"측정 데이터가 '{excel_filename}'에 저장되었습니다.")
    else:
        print("저장할 측정 데이터가 없습니다.")
finally:
    try:
        wb.save(excel_filename)
        wb.close()
        ser.close()
    except Exception as e:
        print(f"종료 중 오류 발생: {e}")
