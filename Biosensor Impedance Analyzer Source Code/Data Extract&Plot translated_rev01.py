import serial
import threading
import queue
import time
import sys
import os
from prompt_toolkit import prompt
from prompt_toolkit.patch_stdout import patch_stdout
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import re
from matplotlib import font_manager, rc
from matplotlib.ticker import ScalarFormatter
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# ------------------------
# 0) Font and Serial Port Settings
# ------------------------
# This path is for a Korean font. It might not be necessary in an English environment.
font_path = 'C:/Windows/Fonts/malgun.ttf' 
try:
    font_prop = font_manager.FontProperties(fname=font_path)
    rc('font', family=font_prop.get_name())
except Exception as e:
    print(f"Font setting error: {e}")
    # Fallback to a default font if the specified font is not found
    rc('font', family='DejaVu Sans') 
plt.rcParams['axes.unicode_minus'] = False

try:
    ser = serial.Serial('COM3', 115200, timeout=0.1)
    time.sleep(0.1)
    ser.reset_input_buffer()
    ser.reset_output_buffer()
    print("Successfully connected to the serial port.")
except serial.SerialException as e:
    print(f"Serial port error: {e}")
    sys.exit(1)

# ------------------------
# 1) Global Variables and Excel Initialization
# ------------------------
prompt_queue = queue.Queue()

calibration_runs = []      # Stores information for each calibration run
calibration_data = []      # Stores calibration data

measurement_data = []      # Accumulates single sweep data (Modes 1, 2, 3) on success
range_data = []            # Accumulates range sweep data (Modes 4, 5) on success

current_mode = None
xAddrStr = ""
yAddrStr = ""
group_selected = None
calibration_impedance = ""
measurement_type = None

currentCoord = None
next_x = None
next_y = None

range_sweep_complete = threading.Event()

save_directory = "C:/Users/Hyunseo/OneDrive/Desktop/Data"
base_filename = "measurement_data"
file_extension = "xlsx"

def get_unique_filename(directory, base_name, extension):
    filename = f"{base_name}.{extension}"
    counter = 2
    while os.path.exists(os.path.join(directory, filename)):
        filename = f"{base_name}_{counter}.{extension}"
        counter += 1
    return os.path.join(directory, filename)

excel_filename = get_unique_filename(save_directory, base_filename, file_extension)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Measurement Data"

current_calibration_run = 0
is_calibrating = False

expected_points = None
actual_count = 0
in_sweep = False
temp_data = []

def initialize_new_calibration_run(run_number):
    global current_calibration_run
    run_number = current_calibration_run
    start_col = 1 + 8 * run_number
    start_col_letter = get_column_letter(start_col)
    start_row = 1
    ws.cell(row=start_row, column=start_col, value="Set Calibration Impedance: ")
    cal_headers = ['Cal Point', 'R / I', '|Z|', 'System Phase']
    for i, header in enumerate(cal_headers):
        ws.cell(row=start_row + 1, column=start_col + i, value=header)
    wb.save(excel_filename)
    print(f"Calibration Run {run_number} initialized. Starting column: {start_col_letter}")
    calibration_runs.append({'run_number': run_number, 'start_col': start_col, 'data': []})

initialize_new_calibration_run(current_calibration_run)
print(f"Excel file '{excel_filename}' has been created.")
sweep_complete = threading.Event()

# ------------------------
# 2) Function to Identify Prompt Lines
# ------------------------
def is_prompt_line(line):
    # These patterns must exactly match the prompts sent from the Arduino.
    prompt_patterns = [
        r"^Enter the start frequency",
        r"^Enter the frequency increment",
        r"^Enter the number of measurements",
        r"^Enter Settling Time Cycles",
        r"^Select Output Excitation Range",
        r"^Select PGA Gain",
        r"^Enter Calibration Impedance",
        r"^Select MUX group",
        r"^X Axis Address",
        r"^Y Axis Address",
        r"^Set AD5933 Mode",
        r"^\s*Bit", # Matches "Bit" with any leading whitespace
        r"^Is this range correct\? \(Y/N\)",
        r"^Enter X-axis increment unit",
        r"^Enter Y-axis increment unit"
    ]
    line_stripped = line.strip()
    for pattern in prompt_patterns:
        if re.search(pattern, line_stripped):
            return True
    return False

# ------------------------
# 3) Data Parsing/Processing Functions
# ------------------------
def parse_calibration_line(line):
    try:
        pattern = r"Cal Point (\d+):\s+R=(-?\d+) / I=(-?\d+)\s+\|Z\|=([\d.]+)\s+System Phase=([\d.+-]+) degrees"
        match = re.match(pattern, line)
        if match:
            cal_point = f"Cal Point {match.group(1)}"
            r_i = f"R={match.group(2)} / I={match.group(3)}"
            z = match.group(4)
            phase = f"{match.group(5)} degrees"
            return [cal_point, r_i, z, phase]
        else:
            return None
    except Exception as e:
        print(f"Calibration data parsing error: {e} - Line: {line}")
        return None

def parse_measurement_line(line):
    # This function is modified to handle "ovf" (overflow) values from the Arduino.
    try:
        # Regex pattern that accepts either a floating point number or the string "ovf"
        value_pattern = r"([-+]?\d+\.\d+|ovf)"
        pattern = (
            r"(\d+\.\d+)kHz:\s+R=(-?\d+)/I=(-?\d+)\s+"
            rf"\|Z\|={value_pattern}\s+"
            rf"Phase=([-+]?\d+\.\d+)\s+degrees\s+" # Phase usually doesn't overflow
            rf"Resistance={value_pattern}\s+"
            rf"Reactance={value_pattern}"
        )
        match = re.match(pattern, line)
        if match:
            freq_khz = float(match.group(1))
            freq = f"{int(freq_khz * 1000)} Hz"
            r_i = f"R={match.group(2)} / I={match.group(3)}"
            
            # Helper function to convert "ovf" to 0.0 or a large number, or parse float
            def parse_value(v_str):
                return 0.0 if v_str == 'ovf' else float(v_str)

            impedance = parse_value(match.group(4))
            phase = float(match.group(5)) # Phase is assumed to be a number
            resistance = parse_value(match.group(6))
            reactance = parse_value(match.group(7))
            
            return [freq, r_i, impedance, phase, resistance, reactance]
        else:
            return None
    except (IndexError, ValueError) as e:
        print(f"Measurement data parsing error: {e} - Line: {line}")
        return None

def add_headers(current_run, headers):
    start_col = current_run['start_col']
    if 'current_row' not in current_run:
        current_run['current_row'] = 1
    header_row = current_run['current_row']
    for i, header in enumerate(headers):
        cell = ws.cell(row=header_row, column=start_col + i, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    current_run['current_row'] += 1
    wb.save(excel_filename)
    print(f"Headers added: {headers}")

HEADER_FIELDS = ['Freq (Hz)', 'R / I', '|Z|', 'Phase (Degrees)', 'Resistance', 'Reactance', 'X', 'Y']

# ------------------------
# 4) Function to Plot Averages and Individual R/I by Frequency
# ------------------------
def plot_average_by_frequency(data, mode_label=""):
    cols = ['freq (Hz)', 'R / I', '|Z|', 'Phase (Degrees)', 'Resistance', 'Reactance', 'X', 'Y']
    df = pd.DataFrame(data, columns=cols)
    if df.empty:
        print("No data to plot.")
        return

    df['Frequency'] = df['freq (Hz)'].apply(lambda x: int(x.replace(' Hz','')))

    for col in ['|Z|', 'Phase (Degrees)', 'Resistance', 'Reactance']:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    df['Coord'] = df.apply(lambda row: f"X={row['X']},Y={row['Y']}", axis=1)
    df_avg = df.groupby('Frequency', as_index=False)[['|Z|', 'Phase (Degrees)', 'Resistance', 'Reactance']].mean()

    def extract_r_i(r_i_str):
        try:
            r, i = r_i_str.split('/')
            r = float(r.replace('R=','').strip())
            i = float(i.replace('I=','').strip())
            return r, i
        except:
            return None, None

    df[['R_val', 'I_val']] = df['R / I'].apply(lambda x: pd.Series(extract_r_i(x)))
    groups = df.groupby('Coord')
    
    color_cycle = plt.colormaps.get_cmap('tab10')
    fig, axs = plt.subplots(2, 3, figsize=(18, 10))

    axs[0, 0].plot(df_avg['Frequency'], df_avg['|Z|'], marker='o', linestyle='-')
    axs[0, 0].set_title('Frequency vs. Average |Z|')
    axs[0, 0].set_xlabel('Frequency (Hz)')
    axs[0, 0].set_ylabel('Average |Z|')
    axs[0, 0].grid(True)

    axs[0, 1].plot(df_avg['Frequency'], df_avg['Phase (Degrees)'], marker='o', linestyle='-')
    axs[0, 1].set_title('Frequency vs. Average Phase')
    axs[0, 1].set_xlabel('Frequency (Hz)')
    axs[0, 1].set_ylabel('Average Phase (Degrees)')
    axs[0, 1].grid(True)

    axs[0, 2].plot(df_avg['Frequency'], df_avg['Resistance'], marker='o', linestyle='-')
    axs[0, 2].set_title('Frequency vs. Average Resistance')
    axs[0, 2].set_xlabel('Frequency (Hz)')
    axs[0, 2].set_ylabel('Average Resistance')
    axs[0, 2].grid(True)

    axs[1, 0].plot(df_avg['Frequency'], df_avg['Reactance'], marker='o', linestyle='-')
    axs[1, 0].set_title('Frequency vs. Average Reactance')
    axs[1, 0].set_xlabel('Frequency (Hz)')
    axs[1, 0].set_ylabel('Average Reactance')
    axs[1, 0].grid(True)

    for idx, (coord, group) in enumerate(groups):
        color = color_cycle(idx % 10)
        axs[1, 1].scatter(group['Frequency'], group['R_val'], marker='o', color=color, label=f'R ({coord})')
        axs[1, 1].scatter(group['Frequency'], group['I_val'], marker='x', color=color, label=f'I ({coord})')
    axs[1, 1].set_title('Frequency vs. Individual R / I')
    axs[1, 1].set_xlabel('Frequency (Hz)')
    axs[1, 1].set_ylabel('R, I values')
    if len(groups) <= 10:
        axs[1, 1].legend()
    axs[1, 1].grid(True)

    axs[1, 2].axis('off')
    
    if mode_label == '2':
        title_label = "Rcal Position"
    elif mode_label == '1':
        title_label = "COB Position"
    elif mode_label == '4':
        title_label = "COB Range Sweep"
    elif mode_label == '5':
        title_label = "COB Range Step Sweep"
    else:
        title_label = "Impedance"
        
    fig.suptitle(f"{title_label} - Averages by Frequency (|Z|, Phase, Resistance, Reactance) and Individual R/I", fontsize=16)
    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    plt.show()

# ------------------------
# 5) Function to Plot Raw Data (Individual)
# ------------------------
def plot_data(data, mode_label=""):
    cols = ['freq (Hz)', 'R / I', '|Z|', 'Phase (Degrees)', 'Resistance', 'Reactance', 'X', 'Y']
    df = pd.DataFrame(data, columns=cols)
    if df.empty:
        print("No data to plot.")
        return
    df['Frequency'] = df['freq (Hz)'].apply(lambda x: int(x.replace(' Hz','')))

    def extract_r_i(r_i_str):
        try:
            r, i = r_i_str.split('/')
            r = float(r.replace('R=','').strip())
            i = float(i.replace('I=','').strip())
            return r, i
        except:
            return None, None

    df[['R', 'I']] = df['R / I'].apply(lambda x: pd.Series(extract_r_i(x)))
    df = df.dropna(subset=['R','I'])
    df['Coord'] = df.apply(lambda row: f"X={row['X']},Y={row['Y']}", axis=1)

    if mode_label == '2':
        suptitle = "Rcal Position Impedance Measurement Results"
    elif mode_label == '1':
        suptitle = "COB Impedance Measurement Results"
    elif mode_label == '4':
        suptitle = "COB Range Sweep Results"
    elif mode_label == '5':
        suptitle = "COB Range Step Sweep Results"
    else:
        suptitle = "Impedance Measurement Results"

    color_cycle = plt.colormaps.get_cmap('tab10')
    groups = df.groupby('Coord')

    plt.figure(figsize=(15, 10))

    ax1 = plt.subplot(2,2,1)
    for idx, (coord, group) in enumerate(groups):
        color = color_cycle(idx % 10)
        ax1.scatter(group['Frequency'], group['R'], marker='o', color=color, label=f'R ({coord})')
        ax1.scatter(group['Frequency'], group['I'], marker='x', color=color, label=f'I ({coord})')
    ax1.set_title('Frequency vs. R and I')
    ax1.set_xlabel('Frequency (Hz)')
    ax1.set_ylabel('R and I')
    if len(groups) <= 10:
        ax1.legend()
    ax1.grid(True)

    ax2 = plt.subplot(2,2,2)
    for idx, (coord, group) in enumerate(groups):
        color = color_cycle(idx % 10)
        ax2.scatter(group['Frequency'], group['|Z|'], marker='s', color=color, label=coord)
    ax2.set_title('Frequency vs. |Z|')
    ax2.set_xlabel('Frequency (Hz)')
    ax2.set_ylabel('|Z| (Ohm)')
    ax2.set_yscale('linear')
    ax2.yaxis.set_major_formatter(ScalarFormatter(useOffset=False))
    ax2.grid(True)
    if len(groups) <= 10:
        ax2.legend()

    ax3 = plt.subplot(2,2,3)
    for idx, (coord, group) in enumerate(groups):
        color = color_cycle(idx % 10)
        ax3.scatter(group['Frequency'], group['Phase (Degrees)'], marker='^', color=color, label=coord)
    ax3.set_title('Frequency vs. Phase (Degrees)')
    ax3.set_xlabel('Frequency (Hz)')
    ax3.set_ylabel('Phase (Degrees)')
    ax3.set_ylim(-180, 180)
    ax3.grid(True)
    if len(groups) <= 10:
        ax3.legend()

    ax4 = plt.subplot(2,2,4)
    for idx, (coord, group) in enumerate(groups):
        color = color_cycle(idx % 10)
        ax4.scatter(group['Frequency'], group['Resistance'], marker='D', color=color, label=f'Res ({coord})')
        ax4.scatter(group['Frequency'], group['Reactance'], marker='v', color=color, label=f'React ({coord})')
    ax4.set_title('Frequency vs. Resistance & Reactance')
    ax4.set_xlabel('Frequency (Hz)')
    ax4.set_ylabel('Value')
    ax4.grid(True)
    if len(groups) <= 10:
        ax4.legend()

    plt.suptitle(suptitle, fontsize=16)
    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    plt.show()

# ------------------------
# 6) Function to Write Temporary Data to Excel (for Range Sweep only)
# ------------------------
def write_temp_data_to_excel(temp_data):
    global measurement_type, measurement_data, range_data
    global calibration_runs

    if not temp_data or not calibration_runs:
        return

    # Most recent calibration_run
    current_run = calibration_runs[-1]
    start_col = current_run['start_col']
    if 'current_row' not in current_run:
        current_run['current_row'] = 3

    # Write to Excel + update range_data (for range sweep modes)
    for row_data in temp_data:
        measurement_row = current_run['current_row']
        for i, datum in enumerate(row_data):
            ws.cell(row=measurement_row, column=start_col + i, value=datum)
        current_run['current_row'] += 1
        if measurement_type in ['COB-range', 'COB-range-step']:
            range_data.append(row_data)

    wb.save(excel_filename)
    print(f"[INFO] Successfully wrote {len(temp_data)} items from temp_data to Excel.")

# ------------------------
# 7) Serial Reception Thread
# ------------------------
def read_from_port():
    global current_mode, xAddrStr, yAddrStr, calibration_impedance, group_selected
    global measurement_type, current_calibration_run, is_calibrating
    global currentCoord, next_x, next_y
    global expected_points, actual_count, in_sweep, temp_data

    while True:
        if not ser.is_open:
            print("Serial port is closed. Exiting data reception loop.")
            break
        try:
            line = ser.readline().decode('utf-8', errors='ignore').strip()
            if line:
                # Handshaking process specifically for range sweep modes
                if measurement_type in ['COB-range', 'COB-range-step']:
                    if line == "SWEEP_START":
                        print("[INFO] SWEEP_START detected -> Initializing temp_data, actual_count=0")
                        temp_data = []
                        actual_count = 0
                        in_sweep = True
                        continue

                    if line == "SWEEP_DONE":
                        print(f"[INFO] SWEEP_DONE detected. actual_count={actual_count} / expected_points={expected_points}")
                        in_sweep = False
                        if expected_points is not None and actual_count == expected_points:
                            print("[INFO] -> Data count matches. Writing temp_data and sending STORE_OK.")
                            write_temp_data_to_excel(temp_data)
                            ser.write(b"STORE_OK\n")
                        else:
                            print("[WARNING] -> Data count mismatch. Discarding temp_data. Not sending STORE_OK (to trigger re-measurement).")
                        continue

                # Existing logic
                if "Starting Calibration." in line:
                    if not is_calibrating:
                        print("\n[INFO] Starting calibration. Initializing a new calibration run.\n")
                        current_calibration_run += 1
                        initialize_new_calibration_run(current_calibration_run)
                        is_calibrating = False
                        ser.reset_input_buffer()
                        ser.reset_output_buffer()
                    continue

                if "ESP-ROM" in line:
                    print("\n[INFO] Device has been reset. Starting a new calibration run.\n")
                    current_calibration_run += 1
                    initialize_new_calibration_run(current_calibration_run)
                    ser.reset_input_buffer()
                    ser.reset_output_buffer()
                    continue

                if is_prompt_line(line):
                    prompt_queue.put(line)
                    continue

                if "[INFO] Set Calibration Impedance" in line:
                    parts = line.split(':')
                    if len(parts) >= 2:
                        calibration_impedance = parts[1].strip().split(' ')[0]
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            ws.cell(row=1, column=start_col,
                                    value=f"Set Calibration Impedance: {calibration_impedance} ohm")
                            wb.save(excel_filename)
                            print(line)
                    continue

                if line.startswith("Cal Point"):
                    cal_data = parse_calibration_line(line)
                    if cal_data and calibration_runs:
                        calibration_data.append(cal_data)
                        current_run = calibration_runs[-1]
                        start_col = current_run['start_col']
                        if 'current_row' not in current_run:
                            current_run['current_row'] = 3
                        for i, data_item in enumerate(cal_data):
                            ws.cell(row=current_run['current_row'], column=start_col + i, value=data_item)
                        current_run['current_row'] += 1
                        wb.save(excel_filename)
                        print("\t".join(map(str, cal_data)))
                    continue

                if "[INFO] Set X-axis Address" in line and "Y-axis Address" in line:
                    parts = line.split(',')
                    if len(parts) >= 2:
                        xAddrStr = parts[0].split(':')[-1].strip()
                        yAddrStr = parts[1].split(':')[-1].strip()
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            if 'current_row' not in current_run:
                                current_run['current_row'] = 3
                            row_num = current_run['current_row']
                            ws.cell(row=row_num, column=start_col, value="Set Coordinates")
                            ws.cell(row=row_num, column=start_col + 1, value=f"X={xAddrStr}")
                            ws.cell(row=row_num, column=start_col + 2, value=f"Y={yAddrStr}")
                            current_run['current_row'] += 1
                            wb.save(excel_filename)
                            print(line)
                            if measurement_type in ['COB', 'Rcal', 'COB-diagonal']:
                                currentCoord = (xAddrStr, yAddrStr)
                    continue

                if "Checking impedance at Rcal position." in line:
                    if calibration_runs:
                        current_run = calibration_runs[-1]
                        start_col = current_run['start_col']
                        if 'current_row' not in current_run:
                            current_run['current_row'] = 1
                        current_run['current_row'] += 1
                        ws.cell(row=current_run['current_row'], column=start_col, value="Checking impedance at Rcal position.")
                        current_run['current_row'] += 1
                        wb.save(excel_filename)
                        print(line)
                        measurement_type = 'Rcal'
                        add_headers(current_run, HEADER_FIELDS)
                    continue

                if "Checking impedance of COB." in line:
                    if calibration_runs:
                        current_run = calibration_runs[-1]
                        start_col = current_run['start_col']
                        if 'current_row' not in current_run:
                            current_run['current_row'] = 1
                        current_run['current_row'] += 1
                        ws.cell(row=current_run['current_row'], column=start_col, value="Checking impedance of COB.")
                        current_run['current_row'] += 1
                        wb.save(excel_filename)
                        print(line)
                        measurement_type = 'COB'
                    continue
                
                if "Starting COB Range Sweep" in line:
                    if calibration_runs:
                        current_run = calibration_runs[-1]
                        start_col = current_run['start_col']
                        if 'current_row' not in current_run:
                            current_run['current_row'] = 3
                        current_run['current_row'] += 1
                        ws.cell(row=current_run['current_row'], column=start_col,
                                value="Starting COB Range Sweep (7-bit input).")
                        current_run['current_row'] += 1
                        wb.save(excel_filename)
                        print(line)
                        measurement_type = 'COB-range'
                    continue

                if "Starting COB Range Step Sweep" in line:
                    if calibration_runs:
                        current_run = calibration_runs[-1]
                        start_col = current_run['start_col']
                        if 'current_row' not in current_run:
                            current_run['current_row'] = 3
                        current_run['current_row'] += 1
                        ws.cell(row=current_run['current_row'], column=start_col, value="Starting COB Range Step Sweep (X/Y increment setting).")
                        current_run['current_row'] += 1
                        wb.save(excel_filename)
                        print(line)
                        measurement_type = 'COB-range-step'
                    continue

                if "[INFO] Group" in line and "selected" in line:
                    match_grp = re.search(r"Group\s+(\d+)\s+selected", line)
                    if match_grp:
                        group_selected = match_grp.group(1)
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            if 'current_row' not in current_run:
                                current_run['current_row'] = 3
                            row_num = current_run['current_row']
                            ws.cell(row=row_num, column=start_col, value=f"Group {group_selected} selected")
                            current_run['current_row'] += 1
                            wb.save(excel_filename)
                            print(line)
                            add_headers(current_run, HEADER_FIELDS)
                    continue

                match_coord = re.search(r"Current_Coord->X=([\d]+),Y=([\d]+)", line)
                if match_coord:
                    next_x = match_coord.group(1)
                    next_y = match_coord.group(2)
                    print(line)
                    if calibration_runs:
                        current_run = calibration_runs[-1]
                        start_col = current_run['start_col']
                        if 'current_row' not in current_run:
                            current_run['current_row'] = 3
                        current_run['current_row'] += 1
                        ws.cell(row=current_run['current_row'], column=start_col, value="Current Coordinates")
                        ws.cell(row=current_run['current_row'], column=start_col + 1, value=f"X={next_x}")
                        ws.cell(row=current_run['current_row'], column=start_col + 2, value=f"Y={next_y}")
                        current_run['current_row'] += 1
                        wb.save(excel_filename)
                        currentCoord = (next_x, next_y)
                        next_x = None
                        next_y = None
                    continue

                if "Frequency sweep complete!" in line:
                    print(line)
                    sweep_complete.set()
                    continue

                if "[INFO] COB range sweep complete" in line:
                    print(line)
                    range_sweep_complete.set()
                    continue

                if "[INFO] COB range step sweep complete" in line:
                    print(line)
                    range_sweep_complete.set()
                    continue

                # ---------------------------
                # Measurement Data Parsing
                # ---------------------------
                print(line)
                parsed = parse_measurement_line(line)
                if parsed:
                    if measurement_type in ['COB-range', 'COB-range-step']:
                        if in_sweep:
                            actual_count += 1
                            if currentCoord:
                                parsed.append(currentCoord[0])
                                parsed.append(currentCoord[1])
                            else:
                                parsed.append("N/A")
                                parsed.append("N/A")
                            temp_data.append(parsed)
                    else:
                        if currentCoord:
                            parsed.append(currentCoord[0])
                            parsed.append(currentCoord[1])
                        else:
                            parsed.append("N/A")
                            parsed.append("N/A")
                        measurement_data.append(parsed)
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            if 'current_row' not in current_run:
                                current_run['current_row'] = 3
                            measurement_row = current_run['current_row']
                            for i, datum in enumerate(parsed):
                                ws.cell(row=measurement_row, column=start_col + i, value=datum)
                            current_run['current_row'] += 1
                            wb.save(excel_filename)

        except serial.SerialException as e:
            # If the device is disconnected, this error is often raised.
            # We will print a clean message and stop the thread.
            print("\n[ERROR] Serial port disconnected. Stopping data reception thread.")
            break # Exit the while loop to terminate the thread
        except Exception as e:
            # Catch other potential errors
            print(f"\n[ERROR] An unexpected error occurred in the reading thread: {e}")
            break

# ------------------------
# 8) Main Loop (Program Entry Point)
# ------------------------
thread = threading.Thread(target=read_from_port, daemon=True)
thread.start()

try:
    while True:
        if not prompt_queue.empty():
            prompt_text = prompt_queue.get()
            with patch_stdout():
                user_input = prompt(prompt_text)
            # When prompted for "Enter the number of measurements", set expected_points
            if "Enter the number of measurements" in prompt_text:
                try:
                    num_increments = int(user_input.strip())
                    if 1 <= num_increments <= 100:
                        expected_points = num_increments + 1
                        print(f"[INFO] numIncrements={num_increments}, expected_points={expected_points}")
                    else:
                        print("[WARNING] Input is not in the range 1-100. The device will prompt for re-entry.")
                except ValueError:
                    print("[WARNING] Failed to convert to integer.")
            try:
                ser.write((user_input.strip() + '\n').encode('utf-8'))
                if "Set AD5933 Mode" in prompt_text:
                    current_mode = user_input.strip()
                    if current_mode == '1':
                        measurement_type = 'COB'
                    elif current_mode == '2':
                        measurement_type = 'Rcal'
                    elif current_mode == '3':
                        measurement_type = 'COB-diagonal'
                    elif current_mode == '4':
                        measurement_type = 'COB-range'
                    elif current_mode == '5':
                        measurement_type = 'COB-range-step'
                    elif current_mode == '0':
                        is_calibrating = True
                        current_calibration_run += 1
                        initialize_new_calibration_run(current_calibration_run)
                        print(f"Switched to Calibration Run {current_calibration_run}.")
                    else:
                        measurement_type = 'Unknown'
            except serial.SerialException as e:
                print(f"Error sending data to serial port: {e}")
                break
        elif sweep_complete.is_set():
            # For single sweep modes (COB, Rcal, COB-diagonal), plot immediately
            if measurement_type in ['COB', 'Rcal', 'COB-diagonal']:
                print("\n[INFO] Single sweep complete - plotting data.\n")
                plot_data(measurement_data, mode_label=current_mode)
                measurement_data.clear()
                currentCoord = None
            sweep_complete.clear()
        elif range_sweep_complete.is_set():
            # For range sweep modes, validate user input before plotting
            if measurement_type in ['COB-range', 'COB-range-step']:
                mode_map = {'COB-range': '4', 'COB-range-step': '5'}
                mode_num = mode_map.get(measurement_type)
                print(f"\n[INFO] {measurement_type} complete. Select plot option.\n")
                
                while True:
                    user_choice = prompt("Plotting options (avg/ind): ").strip().lower()
                    if user_choice == 'avg':
                        plot_average_by_frequency(range_data, mode_label=mode_num)
                        break
                    elif user_choice == 'ind':
                        plot_data(range_data, mode_label=mode_num)
                        break
                    else:
                        print("[ERROR] Invalid input. Please enter 'avg' or 'ind'.")

                range_data.clear()
                currentCoord = None
            range_sweep_complete.clear()
        else:
            time.sleep(0.01)
except KeyboardInterrupt:
    print("\nExiting the program.")
    if measurement_data or range_data:
        data_to_plot = range_data if range_data else measurement_data
        
        # Check if it was a range sweep to ask for plotting options
        if measurement_type in ['COB-range', 'COB-range-step']:
             while True:
                user_choice = prompt("Select plot option before exiting (avg/ind): ").strip().lower()
                if user_choice == 'avg':
                    plot_average_by_frequency(data_to_plot, mode_label=current_mode)
                    break
                elif user_choice == 'ind':
                    plot_data(data_to_plot, mode_label=current_mode)
                    break
                else:
                    print("[ERROR] Invalid input. Please enter 'avg' or 'ind'.")
        else: # For other modes, plot directly
            plot_data(data_to_plot, mode_label=current_mode)
            
        print(f"Measurement data has been saved to '{excel_filename}'.")
    else:
        print("No measurement data to save.")
finally:
    try:
        wb.save(excel_filename)
        wb.close()
        ser.close()
    except Exception as e:
        print(f"Error during exit: {e}")
