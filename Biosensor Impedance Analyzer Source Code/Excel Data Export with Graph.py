import serial
import threading
import queue
import time
import sys
import os
from prompt_toolkit import prompt
from prompt_toolkit.patch_stdout import patch_stdout
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import re
from matplotlib import font_manager, rc
from matplotlib.ticker import ScalarFormatter
from openpyxl.utils import get_column_letter

# 한글 폰트 설정
font_path = 'C:/Windows/Fonts/malgun.ttf'  # 맑은 고딕 폰트 경로
try:
    font_prop = font_manager.FontProperties(fname=font_path)
    rc('font', family=font_prop.get_name())
except Exception as e:
    print(f"폰트 설정 오류: {e}")
    rc('font', family='DejaVu Sans')  # 기본 폰트로 설정

# 음수 기호 제대로 표시
plt.rcParams['axes.unicode_minus'] = False

# 시리얼 포트 설정
try:
    ser = serial.Serial('COM5', 115200, timeout=0.1)
    time.sleep(0.1)  # 시리얼 포트 안정화 시간
    ser.reset_input_buffer()
    ser.reset_output_buffer()
    print(f"시리얼 포트 COM5에 연결되었습니다.")
except serial.SerialException as e:
    print(f"시리얼 포트 오류: {e}")
    sys.exit(1)

# 프롬프트 큐 설정
prompt_queue = queue.Queue()

# 데이터 저장 리스트 및 글로벌 변수 설정
calibration_runs = []  # 각 캘리브레이션 런의 데이터를 저장
calibration_data = []  # 각 캘리브레이션 데이터 저장
measurement_data = []
current_mode = None  # 현재 모드를 트래킹
xAddrStr = ""
yAddrStr = ""
group_selected = None  # 선택된 그룹 번호
calibration_impedance = ""

measurement_type = None  # 현재 측정 유형 ('COB' 또는 'Rcal')

# 데이터 저장 경로 및 파일 이름 설정
save_directory = "C:/Users/Hyunseo/OneDrive/Desktop/Data"
base_filename = "measurement_data"
file_extension = "xlsx"

def get_unique_filename(directory, base_name, extension):
    """
    파일이 존재하면 번호를 추가하여 고유한 파일 이름을 반환
    예: measurement_data.xlsx, measurement_data_2.xlsx, measurement_data_3.xlsx, ...
    """
    filename = f"{base_name}.{extension}"
    counter = 2
    while os.path.exists(os.path.join(directory, filename)):
        filename = f"{base_name}_{counter}.{extension}"
        counter += 1
    return os.path.join(directory, filename)

# 엑셀 파일 초기화
excel_filename = get_unique_filename(save_directory, base_filename, file_extension)

# openpyxl을 사용해 새로운 workbook 생성 및 헤더 추가
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Measurement Data"

# 캘리브레이션 런 관리
current_calibration_run = 0  # 현재 캘리브레이션 런 인덱스
is_calibrating = False  # 캘리브레이션 런 초기화 플래그

def initialize_new_calibration_run(run_number):
    """
    새로운 캘리브레이션 런을 초기화하고 엑셀에 헤더를 추가
    """
    global current_calibration_run
    run_number = current_calibration_run
    # 시작 열 계산 (A=1, I=9, Q=17, Y=25, ...)
    start_col = 1 + 8 * run_number
    start_col_letter = get_column_letter(start_col)
    
    # 캘리브레이션 런의 시작 행 (1)
    start_row = 1

    # "설정된 Calibration Impedance : " 기록
    ws.cell(row=start_row, column=start_col, value="설정된 Calibration Impedance : ")

    # 캘리브레이션 헤더 추가
    cal_headers = ['Cal Point', 'R / I', '|Z|', 'System Phase']
    for i, header in enumerate(cal_headers):
        ws.cell(row=start_row + 1, column=start_col + i, value=header)

    wb.save(excel_filename)
    print(f"캘리브레이션 런 {run_number} 초기화 완료. 시작 열: {start_col_letter}")

    # 새로운 런을 리스트에 추가
    calibration_runs.append({'run_number': run_number, 'start_col': start_col, 'data': []})

# 초기 캘리브레이션 런 초기화
initialize_new_calibration_run(current_calibration_run)
print(f"엑셀 파일 '{excel_filename}'이(가) 생성되었습니다.")

# Frequency sweep 완료 신호를 위한 이벤트 설정
sweep_complete = threading.Event()

def is_prompt_line(line):
    # 실제 입력을 요구하는 프롬프트
    prompts = [
        "시작 주파수를 입력하세요",
        "주파수 증가량을 입력하세요",
        "측정 횟수를 입력하세요",
        "Settling Time Cycles를 입력하세요",
        "Output Excitation Range를 선택하세요",
        "PGA Gain을 선택하세요",
        "Calibration Impedance를 입력하세요",
        "MUX 그룹을 선택하세요",
        "X Axis Address",
        "Y Axis Address",
        "AD5933 모드 설정",
    ]
    return any(line.strip().startswith(prompt) for prompt in prompts)

def parse_calibration_line(line):
    """
    Calibration 데이터를 파싱하여 리스트 형태로 반환
    예시 형식: "Cal Point 0: R=5441 / I=-8309    |Z|=9931.97     System Phase=303.22 degrees"
    """
    try:
        # 정규 표현식을 사용하여 측정 데이터 라인만을 매칭
        pattern = r"Cal Point (\d+):\s+R=(-?\d+) / I=(-?\d+)\s+\|Z\|=([\d.]+)\s+System Phase=([\d.+-]+) degrees"
        match = re.match(pattern, line)
        if match:
            cal_point = f"Cal Point {match.group(1)}"
            r_i = f"R={match.group(2)} / I={match.group(3)}"  # 형식 수정
            z = match.group(4)
            phase = f"{match.group(5)} degrees"
            return [cal_point, r_i, z, phase]
        else:
            # 측정 데이터 형식이 아닌 경우 None 반환
            return None
    except Exception as e:
        # 파싱 오류 시 None 반환
        print(f"Calibration 데이터 파싱 오류: {e} - 라인: {line}")
        return None

def parse_measurement_line(line):
    """
    측정 데이터를 파싱하여 리스트 형태로 반환
    예시 형식: "50.00kHz: R=5440/I=-8309          |Z|=200000.00   Phase=0.00 degrees     Resistance=200000.00    Reactance=0.00"
    """
    try:
        # 정규 표현식을 사용하여 측정 데이터 라인만을 매칭
        pattern = r"(\d+\.\d+)kHz:\s+R=(-?\d+)/I=(-?\d+)\s+\|Z\|=([-+]?\d+\.\d+)\s+Phase=([-+]?\d+\.\d+) degrees\s+Resistance=([-+]?\d+\.\d+)\s+Reactance=([-+]?\d+\.\d+)"
        match = re.match(pattern, line)
        if match:
            freq_khz = float(match.group(1))
            freq = f"{int(freq_khz * 1000)} Hz"
            r_i = f"R={match.group(2)} / I={match.group(3)}"
            impedance = float(match.group(4))
            phase = float(match.group(5))
            resistance = float(match.group(6))
            reactance = float(match.group(7))
            return [freq, r_i, impedance, phase, resistance, reactance]
        else:
            # 측정 데이터 형식이 아닌 경우 None 반환
            return None
    except (IndexError, ValueError) as e:
        # 파싱 오류 시 None 반환
        print(f"Measurement 데이터 파싱 오류: {e} - 라인: {line}")
        return None

def add_headers(current_run, headers):
    """
    주어진 현재 런에 헤더를 추가
    """
    start_col = current_run['start_col']
    for i, header in enumerate(headers):
        ws.cell(row=current_run['current_row'], column=start_col + i, value=header)
    current_run['current_row'] += 1
    wb.save(excel_filename)
    print(f"헤더 추가: {headers}")

HEADER_FIELDS = ['Freq (Hz)', 'R / I', '|Z|', 'Phase (Degrees)', 'Resistance', 'Reactance']

def read_from_port():
    global current_mode, xAddrStr, yAddrStr, calibration_impedance, group_selected, measurement_type, current_calibration_run, is_calibrating
    while True:
        try:
            line = ser.readline().decode('utf-8', errors='ignore').strip()
            if line:
                # 캘리브레이션 시작 문구 인식
                if "캘리브레이션을 시작합니다." in line:
                    if not is_calibrating:
                        print("\n캘리브레이션을 시작합니다. 새로운 캘리브레이션 런을 초기화합니다.\n")
                        current_calibration_run += 1
                        initialize_new_calibration_run(current_calibration_run)
                        is_calibrating = False
                        ser.reset_input_buffer()
                        ser.reset_output_buffer()
                    continue

                # 리셋 메시지 감지
                if "ESP-ROM" in line:
                    print("\n디바이스가 리셋되었습니다. 새로운 캘리브레이션 런을 시작합니다.\n")
                    current_calibration_run += 1
                    initialize_new_calibration_run(current_calibration_run)
                    ser.reset_input_buffer()
                    ser.reset_output_buffer()
                    continue

                # 프롬프트 라인 처리
                if is_prompt_line(line):
                    prompt_queue.put(line)
                elif "설정된 Calibration Impedance" in line:
                    parts = line.split(':')
                    if len(parts) >= 2:
                        calibration_impedance = parts[1].strip().split(' ')[0]
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            ws.cell(row=1, column=start_col, value=f"설정된 Calibration Impedance : {calibration_impedance} ohm")
                            wb.save(excel_filename)
                            print(line)
                elif line.startswith("Cal Point"):
                    cal_data = parse_calibration_line(line)
                    if cal_data:
                        calibration_data.append(cal_data)
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            if 'current_row' not in current_run:
                                current_run['current_row'] = 3
                            for i, data in enumerate(cal_data):
                                ws.cell(row=current_run['current_row'], column=start_col + i, value=data)
                            current_run['current_row'] += 1
                            wb.save(excel_filename)
                            print("\t".join(cal_data))
                elif "설정된 X축 Address" in line and "Y축 Address" in line:
                    parts = line.split(',')
                    if len(parts) >= 2:
                        xAddrStr = parts[0].split(':')[-1].strip()
                        yAddrStr = parts[1].split(':')[-1].strip()
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            row_num = current_run['current_row']
                            ws.cell(row=row_num, column=start_col, value=f"설정된 X축 Address : {xAddrStr}, Y축 Address : {yAddrStr}")
                            current_run['current_row'] += 1
                            wb.save(excel_filename)
                            print(line)
                elif "그룹" in line and "선택" in line:
                    match = re.search(r"그룹\s+(\d+)\s+선택", line)
                    if match:
                        group_selected = match.group(1)
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            row_num = current_run['current_row']
                            ws.cell(row=row_num, column=start_col, value=f"그룹 {group_selected} 선택")
                            current_run['current_row'] += 1
                            wb.save(excel_filename)
                            print(line)
                            add_headers(current_run, HEADER_FIELDS)
                elif "COB의 임피던스를 체크합니다." in line:
                    if calibration_runs:
                        current_run = calibration_runs[-1]
                        start_col = current_run['start_col']
                        current_run['current_row'] += 1  # 빈 행 추가
                        ws.cell(row=current_run['current_row'], column=start_col, value="COB 위치의 임피던스를 체크합니다.")
                        current_run['current_row'] += 1
                        wb.save(excel_filename)
                        print(line)
                        measurement_type = 'COB'
                elif "Rcal 위치의 임피던스를 체크합니다." in line:
                    if calibration_runs:
                        current_run = calibration_runs[-1]
                        start_col = current_run['start_col']
                        current_run['current_row'] += 1  # 빈 행 추가
                        ws.cell(row=current_run['current_row'], column=start_col, value="Rcal 위치의 임피던스를 체크합니다.")
                        current_run['current_row'] += 1
                        wb.save(excel_filename)
                        print(line)
                        measurement_type = 'Rcal'
                        add_headers(current_run, HEADER_FIELDS)
                elif "Frequency sweep complete!" in line:
                    print(line)
                    sweep_complete.set()
                else:
                    print(line)
                    data = parse_measurement_line(line)
                    if data:
                        measurement_data.append(data)
                        if calibration_runs:
                            current_run = calibration_runs[-1]
                            start_col = current_run['start_col']
                            measurement_row = current_run['current_row']
                            if measurement_type in ['COB', 'Rcal']:
                                for i, datum in enumerate(data):
                                    ws.cell(row=measurement_row, column=start_col + i, value=datum)
                                current_run['current_row'] += 1
                                wb.save(excel_filename)
        except serial.SerialException as e:
            print(f"데이터 수신 중 시리얼 포트 오류 발생: {e}")
            time.sleep(0.05)
        except Exception as e:
            print(f"데이터 수신 중 오류 발생: {e}")
            time.sleep(0.01)

def plot_data(data):
    """
    수집된 데이터를 기반으로 4개의 그래프를 플로팅
    """
    df = pd.DataFrame(data, columns=['freq (Hz)', 'R / I', '|Z|', 'Phase (Degrees)', 'Resistance', 'Reactance'])
    
    if df.empty:
        print("플로팅할 데이터가 없습니다.")
        return

    # 1. 'freq (Hz)'에서 'Hz' 제거하고 정수형으로 변환
    df['Frequency'] = df['freq (Hz)'].apply(lambda x: int(x.replace(' Hz','')))
    
    # 2. 'R / I'에서 'R='과 'I=' 제거하고 숫자로 변환하여 'R'과 'I' 컬럼 생성
    def extract_r_i(r_i_str):
        try:
            r, i = r_i_str.split('/')
            r = float(r.replace('R=','').strip())
            i = float(i.replace('I=','').strip())
            return r, i
        except:
            return None, None

    df[['R', 'I']] = df['R / I'].apply(lambda x: pd.Series(extract_r_i(x)))
    
    # 3. 'R'과 'I'가 None인 행 제거
    df = df.dropna(subset=['R', 'I'])

    # 4. 'current_mode'에 따른 measurement_type 결정
    if current_mode == '1':
        measurement_type_local = 'COB'
    elif current_mode == '2':
        measurement_type_local = 'Rcal'
    else:
        measurement_type_local = 'Unknown'

    # 5. measurement_type에 따른 그래프 제목 설정
    if measurement_type_local == 'Rcal':
        suptitle = "Rcal 임피던스 측정 결과"
    elif measurement_type_local == 'COB':
        suptitle = f"COB 임피던스 측정 결과 / 설정된 X축 Address : {xAddrStr}, Y축 Address : {yAddrStr} / 그룹 {group_selected} 선택"
    else:
        suptitle = "임피던스 측정 결과"

    # 6. 그래프 플로팅
    plt.figure(figsize=(15, 10))
    
    plt.subplot(2, 2, 1)
    plt.scatter(df['Frequency'], df['R'], label='R', color='blue', marker='o')
    plt.scatter(df['Frequency'], df['I'], label='I', color='orange', marker='x')
    plt.title('Frequency vs R and I')
    plt.xlabel('Frequency (Hz)')
    plt.ylabel('R and I')
    plt.legend()
    plt.grid(True)
    
    plt.subplot(2, 2, 2)
    plt.scatter(df['Frequency'], df['|Z|'], color='green', marker='s')
    plt.title('Frequency vs |Z|')
    plt.xlabel('Frequency (Hz)')
    plt.ylabel('|Z| (Ohm)')

    # Y축을 선형으로 설정
    plt.yscale('linear')
    
    # ScalarFormatter 설정 / 임피던스 Y축 레이블 설정 오류로 도입
    ax = plt.gca()
    formatter = ScalarFormatter(useOffset=False)
    ax.yaxis.set_major_formatter(formatter)
    
    # # Y축 범위 설정(플로팅 범위 고정)
    # z_min = df['|Z|'].min()
    # z_max = df['|Z|'].max()
    # plt.ylim(z_min * 0.9, z_max * 1.1)

    plt.grid(True)
    
    plt.subplot(2, 2, 3)
    plt.scatter(df['Frequency'], df['Phase (Degrees)'], color='red', marker='^')
    plt.title('Frequency vs Phase (Degrees)')
    plt.xlabel('Frequency (Hz)')
    plt.ylabel('Phase (Degrees)')
    plt.grid(True)
    
    plt.subplot(2, 2, 4)
    plt.scatter(df['Frequency'], df['Resistance'], label='Resistance', color='purple', marker='D')
    plt.scatter(df['Frequency'], df['Reactance'], label='Reactance', color='cyan', marker='v')
    plt.title('Frequency vs Resistance & Reactance')
    plt.xlabel('Frequency (Hz)')
    plt.ylabel('Value')
    plt.legend()
    plt.grid(True)
    
    # 제목 설정
    plt.suptitle(suptitle, fontsize=16)
    # 제목 수용을 위한 레이아웃 조절
    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    plt.show()

# 스레드 시작
thread = threading.Thread(target=read_from_port, daemon=True)
thread.start()

# 메인 스레드에서 입력 처리 및 플로팅
# TODO : 입력 프롬프트에서 입력을 받는 중일때 보드 RESET 발생 시, 입력 프롬프트가 초기화되지 않는 문제 존재
try:
    while True:
        if not prompt_queue.empty():
            prompt_text = prompt_queue.get()
            with patch_stdout():
                user_input = prompt(prompt_text + ' ')
            # 사용자 입력을 시리얼 포트로 전송
            try:
                ser.write((user_input.strip() + '\n').encode('utf-8'))
                # 입력에 따라 현재 모드 업데이트
                if "AD5933 모드 설정" in prompt_text:
                    current_mode = user_input.strip()
                    # 측정 유형 설정
                    if current_mode == '1':
                        measurement_type = 'COB'
                    elif current_mode == '2':
                        measurement_type = 'Rcal'
                    elif current_mode == '0':
                        # 0번 선택 시 다음 캘리브레이션 런으로 이동
                        is_calibrating = True  # 캘리브레이션 런 초기화 플래그 설정
                        current_calibration_run += 1
                        initialize_new_calibration_run(current_calibration_run)
                        print(f"캘리브레이션 런 {current_calibration_run}으로 이동했습니다.")
                    else:
                        measurement_type = 'Unknown'
            except serial.SerialException as e:
                print(f"시리얼 포트로 데이터 전송 중 오류 발생: {e}")
                break
        elif sweep_complete.is_set():
            # Sweep 완료 시 현재 measurement_data를 복사하여 플로팅
            data_to_plot = measurement_data.copy()
            plot_data(data_to_plot)
            # measurement_data 초기화
            measurement_data.clear()
            sweep_complete.clear()
        else:
            time.sleep(0.01)
except KeyboardInterrupt:
    print("\n프로그램을 종료합니다.")
    # 프로그램 종료 시 그래프 플로팅
    if measurement_data:
        plot_data(measurement_data)
        print(f"측정 데이터가 '{excel_filename}'에 저장되었습니다.")
    else:
        print("저장할 측정 데이터가 없습니다.")
finally:
    try:
        wb.save(excel_filename)
        wb.close()
        ser.close()
    except Exception as e:
        print(f"종료 중 오류 발생: {e}")